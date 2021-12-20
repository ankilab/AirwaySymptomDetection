import xlsxwriter
import os
import pandas as pd
import numpy as np
import tensorflow as tf
from openpyxl import load_workbook


class GeneticAlgorithm:
    """
    Class to run a genetic algorithm and create everything around it, e.g., an Excel-file with all results.
    """
    def __init__(self, path_to_file, nb_epochs, force_new_init=False):
        """
        Constructor for class GeneticAlgorithm.
        Args:
            path_to_file: Path where to save and filename of the excel-file.
            nb_epochs: Number of how many epochs the neural network will be trained.
            force_new_init: If True an existing excel-file in the given directory with the same filename
                            will be overwritten.
        """
        self.path = path_to_file
        self.nb_conv_layer = 5
        self.nb_conv_filter = [8, 16, 32, 64, 128]
        self.size_conv_filter = [3]
        self.nb_epochs = nb_epochs

        self.writer = None  # used for writing into the excel workbook

        if os.path.exists(self.path) is False or force_new_init is True:
            self.__init_excel()

    def __init_excel(self):
        """
        Initializes an excel-file where to save the genepool and all results.
        """
        workbook = xlsxwriter.Workbook(self.path)
        worksheet_genomes = workbook.add_worksheet("Genome")
        worksheet_genepool = workbook.add_worksheet("Genpool")
        workbook.add_worksheet("Overview")
        workbook.add_worksheet("Training_Detailed")

        self.__init_header_genome(worksheet_genomes)
        self.__init_header_genepool(worksheet_genepool)
        self.__init_genpool(worksheet_genepool)
        workbook.close()

        book = load_workbook(self.path)
        self.writer = pd.ExcelWriter(self.path, engine='openpyxl')
        self.writer.book = book
        self.writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        self.__init_training_detailed_sheet()
        self.__init_overview()

    def get_tf_model(self, model_parameters, input_shape):
        """
        Converts a DataFrame with model-parameters into a trainable tensorflow-model.
        """
        nb_conv_layer = model_parameters["Number_Conv_Layer"][0]
        batch_norm = model_parameters["Batch_Norm"][0]
        activation = model_parameters["Activation"][0]

        # get correct activation function as TensorFlow-object
        if activation == 'relu':
            activation_obj = tf.nn.relu
        elif activation == 'relu6':
            activation_obj = tf.nn.relu6
        else:
            activation_obj = tf.nn.leaky_relu

        # whether to use residual connections or not
        residual_connections = model_parameters["Residual_Connections"][0]

        # extract some information for each layer
        nb_conv_filter = []
        filter_sizes = []
        max_pools = []
        for i in range(1, nb_conv_layer + 1):
            nb_conv_filter.append(int(model_parameters["Number_Conv_Filter_" + str(i)][0]))
            filter_sizes.append(int(model_parameters["Filter_Size_" + str(i)][0]))
            max_pools.append(model_parameters["Max_Pool_" + str(i)][0])

        # create our tf-model and add all layers
        input_layer = tf.keras.Input(input_shape)
        x = input_layer
        for i in range(nb_conv_layer):
            downsample = tf.keras.Sequential()
            downsample.add(tf.keras.layers.Conv2D(filters=nb_conv_filter[i],
                                                  kernel_size=(1, 1),
                                                  strides=1))
            residual = downsample(x)

            x = tf.keras.layers.Conv2D(filters=nb_conv_filter[i], kernel_size=filter_sizes[i], padding='same')(x)

            if bool(batch_norm) is True:
                x = tf.keras.layers.BatchNormalization()(x)

            if bool(residual_connections) is True:
                x = tf.keras.layers.Add()([residual, x])

            x = activation_obj(x)

            if bool(max_pools[i]) is True:
                x = tf.keras.layers.MaxPooling2D()(x)

        x = tf.keras.layers.GlobalAveragePooling2D()(x)
        output = tf.keras.layers.Dense(4, activation='softmax')(x)

        model = tf.keras.Model(input_layer, output)
        return model

    @staticmethod
    def __get_random_value(df):
        """
        Gets a random value from a DataFrame-column where the value is != -1
        e.g. one random value from df=[8, 16, 32, 64, 128, -1, -1, -1,...].
        """
        df = np.array(df)
        df = np.random.choice(df[df != -1])
        return df

    def create_random_model_parameters_from_genpool(self):
        """
        Loads all possible parameters from the genepool and creates a random model from those.
        """
        df = pd.read_excel(self.path, sheet_name="Genpool", engine='openpyxl').fillna(-1)

        # extract random parameters from our genpool
        random_nb_conv_layer = self.__get_random_value(df['Number_Conv_Layer'])
        residual_connections = self.__get_random_value(df['Residual_Connections'])
        batch_norm = self.__get_random_value(df['Batch_Norm'])
        activation = self.__get_random_value(df['Activation'])

        # create arrays to save the parameters which are used inside the for-loop
        max_nb_conv_layer = np.array(df['Number_Conv_Layer'])
        max_nb_conv_layer = np.max(max_nb_conv_layer[max_nb_conv_layer != -1])
        nb_conv_filters = np.zeros(max_nb_conv_layer)
        kernel_sizes = np.zeros(max_nb_conv_layer)
        max_pools = []

        for i in range(max_nb_conv_layer):
            filters = self.__get_random_value(df['Number_Conv_Filter_'+str(i+1)])
            kernel_size = self.__get_random_value(df['Filter_Size_'+str(i+1)])
            max_pool = self.__get_random_value(df['Max_Pool_'+str(i+1)])

            # save each parameter for each for-loop-iteration to save it later in "model_parameters"
            nb_conv_filters[i] = filters
            kernel_sizes[i] = kernel_size
            max_pools.append(max_pool)

        # save all used parameters into a dataframe to save them in our excel file later
        model_parameters = self.__get_model_as_data_frame(random_nb_conv_layer, nb_conv_filters, kernel_sizes,
                                                          residual_connections, batch_norm, max_pools, activation)

        return model_parameters

    def crossover_best_models(self, models_summaries, nb_new_models, n_best_models, prob_mutation=5):
        """
        Crossover the n_best_models into nb_new_models new models using the Fitness as criteria.
        """
        models_summaries = pd.DataFrame(models_summaries)
        models_summaries = models_summaries.sort_values('Fitness', ascending=False)
        best_models = models_summaries[0:n_best_models]

        # load genpool for random mutations
        df_genpool = pd.read_excel(self.path, sheet_name="Genpool", engine='openpyxl').fillna(-1)

        new_models = []
        # create new models
        for i in range(nb_new_models):
            nb_conv_layer = np.random.choice(best_models["Number_Conv_Layer"])
            if np.random.randint(0, 100) < prob_mutation:
                # if mutation select a random other value (make sure to not select the previous value)
                _df = df_genpool["Number_Conv_Layer"]
                _nb_conv_layer = self.__get_random_value(_df.loc[(_df != nb_conv_layer) & (_df != 1)])

                print(f"Mutation: {_nb_conv_layer} Conv Layer instead of {nb_conv_layer}")
                nb_conv_layer = _nb_conv_layer

            max_nb_conv_layer = np.array(df_genpool['Number_Conv_Layer'])
            max_nb_conv_layer = np.max(max_nb_conv_layer[max_nb_conv_layer != -1])
            nb_conv_filters = []
            kernel_sizes = []
            max_pools = []
            for j in range(1, max_nb_conv_layer+1):
                nb_conv_filter = np.random.choice(best_models["Number_Conv_Filter_" + str(j)])
                if np.random.randint(0, 100) < prob_mutation:
                    _df = df_genpool["Number_Conv_Filter_" + str(j)]
                    _nb_conv_filter = self.__get_random_value(_df.loc[_df != nb_conv_filter])

                    print(f"Mutation: {_nb_conv_filter} Conv Filter instead of {nb_conv_filter}")
                    nb_conv_filter = _nb_conv_filter

                kernel_size = np.random.choice(best_models["Filter_Size_" + str(j)])
                if np.random.randint(0, 100) < prob_mutation:
                    _df = df_genpool["Filter_Size_" + str(j)]
                    if len(_df.loc[_df != -1]) > 1:
                        _kernel_size = self.__get_random_value(_df.loc[_df != kernel_size])

                        print(f"Mutation: {_kernel_size} Kernel Size instead of {kernel_size}")
                        kernel_size = _kernel_size

                max_pool = np.random.choice(best_models["Max_Pool_" + str(j)])
                if np.random.randint(0, 100) < prob_mutation:
                    _df = df_genpool["Max_Pool_" + str(j)]
                    _max_pool = self.__get_random_value(_df.loc[_df != max_pool])

                    print(f"Mutation: {_max_pool} Max Pool instead of {max_pool}")
                    max_pool = _max_pool

                nb_conv_filters.append(int(nb_conv_filter))
                kernel_sizes.append(int(kernel_size))
                max_pools.append(max_pool)

            residual_connections = np.random.choice(best_models["Residual_Connections"])
            if np.random.randint(0, 100) < prob_mutation:
                _df = df_genpool["Residual_Connections"]
                _residual_connections = self.__get_random_value(_df.loc[_df != residual_connections])

                print(f"Mutation: {_residual_connections} Residual Connections instead of {residual_connections}")
                residual_connections = _residual_connections

            batch_norm = np.random.choice(best_models["Batch_Norm"])
            if np.random.randint(0, 100) < prob_mutation:
                _df = df_genpool["Batch_Norm"]
                _batch_norm = self.__get_random_value(_df.loc[_df != batch_norm])

                print(f"Mutation: {_batch_norm} Batch Norm instead of {batch_norm}")
                batch_norm = _batch_norm

            activation = np.random.choice(best_models["Activation"])
            if np.random.randint(0, 100) < prob_mutation:
                _df = df_genpool["Activation"]
                _activation = self.__get_random_value(_df.loc[_df != activation])

                print(f"Mutation: {_activation} Activation instead of {activation}")
                activation = _activation

            model_parameters = self.__get_model_as_data_frame(nb_conv_layer, nb_conv_filters, kernel_sizes,
                                                              residual_connections, batch_norm, max_pools, activation)

            new_models.append(model_parameters)

        return new_models

    @staticmethod
    def __get_model_as_data_frame(nb_conv_layer, nb_conv_filters, kernel_sizes, residual_connections, batch_norm, max_pools, activation):
        """
        Converts all model parameters into one single dataframe.
        """
        model_parameters = pd.DataFrame({"Number_Conv_Layer": nb_conv_layer}, index=[0])
        for idx, conv_filter in enumerate(nb_conv_filters):
            model_parameters = pd.concat(
                [model_parameters, pd.DataFrame({"Number_Conv_Filter_" + str(idx+1): conv_filter}, index=[0])], axis=1)

        for idx, kernel_size in enumerate(kernel_sizes):
            model_parameters = pd.concat(
                [model_parameters, pd.DataFrame({"Filter_Size_" + str(idx + 1): kernel_size}, index=[0])], axis=1)

        model_parameters = pd.concat([model_parameters, pd.DataFrame({"Residual_Connections": residual_connections}, index=[0])], axis=1)
        model_parameters = pd.concat([model_parameters, pd.DataFrame({"Batch_Norm": batch_norm}, index=[0])], axis=1)

        for idx, max_pool in enumerate(max_pools):
            model_parameters = pd.concat(
                [model_parameters, pd.DataFrame({"Max_Pool_" + str(idx + 1): max_pool}, index=[0])], axis=1)

        model_parameters = pd.concat([model_parameters, pd.DataFrame({"Activation": activation}, index=[0])], axis=1)

        return model_parameters

    def save_model(self, model_summary):
        """
        Saves all used model parameters as "Genome" to the excel-file.
        """
        df = pd.read_excel(self.path, sheet_name="Genome", engine='openpyxl')
        df = pd.concat([df, model_summary], axis=0)
        df.to_excel(self.writer, "Genome", index=False)
        self.writer.save()

    def save_training_results(self, training_results):
        """
        Saves train-loss, train-acc, val-loss and val-acc to the excel-file.
        """
        df = pd.read_excel(self.path, sheet_name="Training_Detailed", engine='openpyxl')
        df.loc[len(df)] = training_results
        df.to_excel(self.writer, "Training_Detailed", index=False)
        self.writer.save()

    def save_overview(self, generation, val_accs, fitnesses):
        """
        Saves a short overview of a whole generation to the excel-file.
        """
        df = pd.read_excel(self.path, sheet_name="Overview", engine='openpyxl')
        df_new = pd.DataFrame({"Generation": generation,
                               "Mean_Val_Acc": np.mean(val_accs),
                               "Top_Val_Acc": np.max(val_accs),
                               "Mean_Fitness": np.mean(fitnesses),
                               "Top_Fitness": np.max(fitnesses)}, index=[0])
        df = pd.concat([df, df_new], axis=0)
        df.to_excel(self.writer, "Overview", index=False)
        self.writer.save()

    def __init_header_genome(self, worksheet):
        """
        Initializes the header of worksheet "Genome" in the excel-file.
        """
        start = 'A'
        if worksheet.name == 'Genome':
            worksheet.write(start + str(1), "Index")
            start = increment_char(start)
            worksheet.write(start + str(1), "Generation/Individuum")
            start = increment_char(start)

        worksheet.write(start + str(1), "Number_Conv_Layer")
        start = chr(ord(start) + 1)
        for i in range(self.nb_conv_layer):
            worksheet.write(start + str(1), "Number_Conv_Filter_" + str(i + 1))
            start = increment_char(start)

        for i in range(self.nb_conv_layer):
            worksheet.write(start + str(1), "Filter_Size_" + str(i + 1))
            start = increment_char(start)

        worksheet.write(start + str(1), "Residual_Connections")
        start = increment_char(start)

        worksheet.write(start + str(1), "Batch_Norm")
        start = increment_char(start)

        for i in range(self.nb_conv_layer):
            worksheet.write(start + str(1), "Max_Pool_" + str(i + 1))
            start = increment_char(start)

        worksheet.write(start + str(1), "Activation")
        start = increment_char(start)
        worksheet.write(start + str(1), "Train_Acc")
        start = increment_char(start)
        worksheet.write(start + str(1), "Val_Acc")
        start = increment_char(start)
        worksheet.write(start + str(1), "Inference_Time")
        start = increment_char(start)
        worksheet.write(start + str(1), "Fitness")

    def __init_header_genepool(self, worksheet):
        """
        Initializes the header of worksheet "Genepool" in the excel-file.
        """
        start = 'A'
        worksheet.write(start + str(1), "Number_Conv_Layer")
        start = increment_char(start)
        for i in range(self.nb_conv_layer):
            worksheet.write(start + str(1), "Number_Conv_Filter_" + str(i + 1))
            start = increment_char(start)

        for i in range(self.nb_conv_layer):
            worksheet.write(start + str(1), "Filter_Size_" + str(i + 1))
            start = increment_char(start)

        for i in range(self.nb_conv_layer):
            worksheet.write(start + str(1), "Max_Pool_" + str(i + 1))
            start = increment_char(start)

        worksheet.write(start + str(1), "Residual_Connections")
        start = increment_char(start)
        worksheet.write(start + str(1), "Batch_Norm")
        start = increment_char(start)
        worksheet.write(start + str(1), "Activation")

    def __init_genpool(self, worksheet):
        """
        Initializes all values of the Genepool specified in the constructor of this class.
        """
        start = 'B'
        for i in range(self.nb_conv_layer):
            worksheet.write('A' + str(i + 2), str(i+1))

            for idx, nb_cf in enumerate(self.nb_conv_filter):
                worksheet.write(start + str(idx + 2), str(nb_cf))
            start = increment_char(start)

        for i in range(self.nb_conv_layer):
            for idx, f_s in enumerate(self.size_conv_filter):
                worksheet.write(start + str(idx + 2), str(f_s))
            start = increment_char(start)

        # possible values for residual connections
        worksheet.write(start + str(2), "TRUE")
        worksheet.write(start + str(3), "FALSE")
        start = increment_char(start)

        # possible values for batch_norm
        worksheet.write(start + str(2), "TRUE")
        worksheet.write(start + str(3), "FALSE")
        start = increment_char(start)

        # possible values for max_pool
        for i in range(self.nb_conv_layer):
            worksheet.write(start + str(2), "TRUE")
            worksheet.write(start + str(3), "FALSE")
            start = increment_char(start)

        # possible values for conv activation
        worksheet.write(start + str(2), "relu")
        worksheet.write(start + str(3), "relu6")
        worksheet.write(start + str(4), "leaky_relu")

    def __init_training_detailed_sheet(self):
        """
        Initializes the worksheet "Training_Detailed" in the excel-file.
        """
        df_1 = pd.DataFrame()
        df_1["Index"] = np.nan
        df_2 = pd.DataFrame()
        df_3 = pd.DataFrame()
        df_4 = pd.DataFrame()
        for i in range(self.nb_epochs):
            df_1["Train_Loss_" + str(i)] = np.nan
            df_2["Val_Loss_" + str(i)] = np.nan
            df_3["Train_Acc_" + str(i)] = np.nan
            df_4["Val_Acc_" + str(i)] = np.nan

        df_final = df_1.join(df_2)
        df_final = df_final.join(df_3)
        df_final = df_final.join(df_4)

        df_final.to_excel(self.writer, "Training_Detailed", index=False)
        self.writer.save()

    def __init_overview(self):
        """
        Initializes the worksheet "Overview" in the excel-file.
        """
        df = pd.DataFrame()
        df["Generation"] = np.nan
        df["Mean_Val_Acc"] = np.nan
        df["Top_Val_Acc"] = np.nan
        df["Mean_Fitness"] = np.nan
        df["Top_Fitness"] = np.nan
        df.to_excel(self.writer, "Overview", index=False)
        self.writer.save()


def increment_char(char):
    """
    Increments a character by one. Example: 'C' --> 'D'.
    """
    return chr(ord(char) + 1)

